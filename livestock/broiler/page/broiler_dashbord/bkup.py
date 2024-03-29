import frappe
import erpnext
from frappe.utils import getdate,add_days,get_first_day,get_last_day,nowdate,flt,date_diff
from erpnext.stock.get_item_details import (
	get_conversion_factor,
	get_default_cost_center,
)
from erpnext.stock.utils import (get_incoming_rate)
from erpnext.stock.doctype.item.item import get_item_defaults 

@frappe.whitelist()
def get_company_list():
    data = {}
    data["companys"] = frappe.get_list("Company", fields=['name'],limit_page_length=0, order_by="name",debug=0)
    return data

@frappe.whitelist()
def get_batch_list(company=None):
    data = {}
    data["batchs"] = frappe.get_all("Broiler Batch",filters={'docstatus':['!=','2'],'company':company},fields=['name'],limit_page_length=0, order_by="name",debug=0)
    return data

@frappe.whitelist()
def get_sale_item_list():
    itemgp=['CHICKEN PRODUCTS - ACACIA','CHICKEN PRODUCTS - AL FAKHER','CHICKEN PRODUCTS - AUH']
    return frappe.db.get_all('Item',filters={'item_group':['in',itemgp]},fields=['item_code'],pluck='item_code')

def get_item_price(company,item,warehouse,date,qty,uom):
    #item_account_details = get_item_defaults(item, company)
	#stock_uom = item_account_details.stock_uom
    conversion_factor = get_conversion_factor(item,uom).get("conversion_factor") or 1
    base_rate= get_incoming_rate({
						"item_code": item,
						"warehouse": warehouse,
						"posting_date": date,
						"posting_time": '23:59:59',
						"qty": -1 * qty,
                        'company':company
					}) 
    return base_rate*conversion_factor
@frappe.whitelist()
def get_report(company,batch):
    layer=frappe.get_doc('Broiler Batch',batch)
    default_currency=frappe.get_value('Company',layer.company,'default_currency') 
    
    pjt=frappe.db.get_value('Project',layer.project,['expected_start_date','expected_end_date'], as_dict=1)
   
    end_date=layer.end_date
    if pjt:
        if not layer.end_date:
            end_date=getdate(pjt.expected_end_date) or nowdate()

    rear_perid=[]
    
    doc_placed_date=getdate(layer.start_date)
    rear_start_date=doc_placed_date
    rear_end_date=end_date
    rstart=rear_start_date
    rend=''
    dept=[]
    depts=frappe.get_doc('Broiler Wage Expense Departments',layer.company)
    if depts:
        for dep in depts.department:
            dept.append(dep.name)
    #lay_days=date_diff(lay_end_date,lay_start_date)+1
    
    #----- rearing 
    rear_lbl=['Expense']
    rear_doc=['Doc']
    rear_feed=['Feed']
    rear_vaccine=['Vaccine']
    rear_medicine=['Medicine-Disinfectants']
    #rear_bio=['Biosecurity Requirements']
    #rear_miscel=['Miscellaneous Production Req.']
    rear_other=['Others']
    rear_wages=['Wages']
    reat_tot=['Total']
    rear_doc_tot=0
    rear_feed_tot=0
    rear_vaccine_tot=0
    rear_medicine_tot=0
    rear_bio_tot=0
    rear_miscel_tot=0
    rear_other_tot=0
    rear_wages_tot=0
    reat_tot_tot=0
    rear_ind_expanse=[]
    rear_ind_expanse_tot=0
    rear_ind_itm_cnt=0
    
    
    rear_html=''

    vaccine_list=[]
    itmgroup=frappe.db.get_all('Vaccine Item Group',fields=['item_group'],pluck='item_group')
    for gp in itmgroup:
        itms=frappe.db.get_all('Item',filters={'item_group':gp},fields=['item_code'],pluck='item_code')
        for it in itms:
            vaccine_list.append(it)
    
#=========================================================================
    
    
    date_lbl=rear_start_date.strftime("%d/%m/%y")+'-'+rear_end_date.strftime("%d/%m/%y")
    rear_lbl.append(date_lbl)
    col_tot=0
        #---------------------------------
    base_row_material=''
    finished_product=''
    raw_warehouse=''
    feed_warehouse=''
    if layer.broiler_shed:
        broiler_shed=frappe.db.get_value('Broiler Shed',layer.broiler_shed,['base_row_material','row_material_target_warehouse','product_target_warehouse','product','feed_warehouse'],as_dict=1)
        if broiler_shed:
            base_row_material=broiler_shed.base_row_material
            finished_product=broiler_shed.product
            raw_warehouse=broiler_shed.row_material_target_warehouse
            feed_warehouse=broiler_shed.feed_warehouse

    fromdate=add_days(getdate(rear_start_date),-15)
    if base_row_material:
        itemamt=frappe.db.sql("""select IFNULL(sum(i.net_amount), 0) as amount from `tabPurchase Invoice Item` i left join `tabPurchase Invoice` p on p.name=i.parent 
        where p.posting_date between '{0}' and '{1}' and i.item_code ='{2}' and p.docstatus=1 and i.warehouse='{3}' """.format(fromdate,rear_end_date,base_row_material,raw_warehouse),as_dict=1,debug=0)
            
        if itemamt and itemamt[0].amount < 1:
            itemamt=frappe.db.sql("""select IFNULL(sum(i.net_amount), 0) as amount from `tabPurchase Receipt Item` i left join `tabPurchase Receipt` p on p.name=i.parent 
        where p.posting_date between '{0}' and '{1}' and i.item_code ='{2}' and p.docstatus=1 and i.warehouse='{3}' """.format(fromdate,rear_end_date,base_row_material,raw_warehouse),as_dict=1,debug=0)

        if itemamt and itemamt[0].amount < 1:
            itemamt=frappe.db.sql("""select IFNULL(sum(i.amount), 0) as amount from `tabStock Entry Detail` i left join `tabStock Entry` p on p.name=i.parent 
        where p.stock_entry_type='Material Transfer' and p.posting_date between '{0}' and '{1}' and i.item_code ='{2}' and p.docstatus=1 and i.t_warehouse='{3}' """.format(fromdate,rear_end_date,base_row_material,raw_warehouse),as_dict=1,debug=0)
            
        if itemamt and itemamt[0].amount > 0:
            rear_doc.append(itemamt[0].amount)
            col_tot+=float(itemamt[0].amount)
            rear_doc_tot+=float(itemamt[0].amount)
        else:
            rear_doc.append('0')
    else:
        rear_doc.append('0')
       
       
        #----------------------------------
        
    if layer.feed:
        amt=0
        for itm in layer.feed:
                
            if itm.starter_item:
                rate=get_item_price(layer.company,itm.starter_item,feed_warehouse,itm.date,itm.starter_qty,itm.starter_uom)
                amt+=itm.starter_qty*rate
                       
            if itm.finisher_item:
                rate=get_item_price(layer.company,itm.finisher_item,feed_warehouse,itm.date,itm.finisher_qty,itm.finisher_uom)
                amt+=itm.finisher_qty*rate
                        


        rear_feed.append(amt)
        rear_feed_tot+=amt
        col_tot+=amt
    else:
        rear_feed.append(0)

            #---------------------------------------         
    if layer.medicine:
                
        amt=0
        for itm in layer.medicine:
            rate=get_item_price(layer.company,itm.item,base_row_material,itm.date,itm.qty,itm.uom)
            amt+=itm.qty*rate
                
        rear_medicine.append(amt)
        rear_medicine_tot+=amt
        col_tot+=amt
    else:
        rear_medicine.append(0)

    if layer.vaccine:
                
        amt=0
        for itm in layer.vaccine:
            rate=get_item_price(layer.company,itm.item,base_row_material,itm.date,itm.qty,itm.uom)
            amt+=itm.qty*rate

        rear_vaccine.append(amt)
        rear_vaccine_tot+=amt
        col_tot+=amt
    else:
        rear_vaccine.append(0)
                
            #------------------------------------------
    if layer.used_items:
        amt=0
        for itm in layer.used_items:
            rate=get_item_price(layer.company,itm.item_code,base_row_material,itm.date,itm.qty,itm.uom)
            amt+=itm.qty*rate

        rear_other.append(amt)
        rear_other_tot+=amt
        col_tot+=amt
    else:
        rear_other.append(0)
            
        
            

        #-----------------------------------------
    start=rear_start_date
    end=rear_end_date
    sal=frappe.db.get_all('Salary Slip',filters={'status':['in',['Draft','Submitted']],'company':layer.company,'end_date':['between',[start,end]]},fields=['net_pay'],pluck='net_pay')
    salary=0
    salary_expanse=0
    if sal:
        salary+=sum(c for c in sal)

    wageper=0
    mtotsrt=0
    mtotend=0
    mortmonthavg=0
    live=0
    if layer.daily_mortality:
        for mor in layer.daily_mortality:
            if getdate(mor.date)< getdate(start):
                mtotsrt+=float(mor.total)
            if getdate(start) <= getdate(mor.date)<=getdate(end):
                mtotend+=float(mor.total)

        if mtotend:
            mortmonthavg=float(mtotend)/2
    live=float(layer.doc_placed)-float(mtotsrt)-float(mortmonthavg)
    mort_to=add_days(getdate(start),15)
    totbfor=frappe.db.sql("""select IFNULL(sum(m.total), 0) as tot,b.doc_placed,b.name from `tabBroiler Batch` b 
            left join  `tabMortality` m on b.name=m.parent and m.date < '{1}'            
            where b.company='{0}' 
            and ((b.receiving_date < '{2}' and (b.end_date is NULL or b.end_date='')) 
            or (b.receiving_date < '{2}' and b.end_date >'{1}')) and b.name<>'{3}'
            group by b.name""".format(layer.company,start,end,layer.name),as_dict=1,debug=0)
            
    totcurrent=frappe.db.sql("""select IFNULL(sum(m.total), 0) as tot,b.name from `tabBroiler Batch` b 
        left join `tabMortality` m on b.name=m.parent and m.date between '{1}' and '{2}' where
            b.company='{0}' and b.name<>'{3}' group by b.name""".format(layer.company,start,end,layer.name),as_dict=1,debug=0)
    crr={}
    if totcurrent:
        for cr in totcurrent:
            crr.update({cr.name:cr.tot})
    totlive=0
    if totbfor:
        for bf in totbfor:
            totavg=0
            if crr.get(bf.name):
                totavg=float(crr.get(bf.name))/2

            totlive+=bf.doc_placed-bf.tot-totavg

    if totlive:
        wageper=(float(live)*100)/float(totlive)
    else:
        wageper=100

    if salary:
            
        if wageper:
            salary_expanse=(float(salary)/100)*float(wageper)                
        else:
            batchcount=frappe.db.sql(""" select count(name) as cnt from `tabBroiler Batch` where  status='Open' """,as_dict=1,debug=0)
            if batchcount:
                salary_expanse=float(salary)/float(batchcount[0].cnt)

        rear_wages.append(salary_expanse)
        rear_wages_tot+=salary_expanse
        col_tot+=salary_expanse
    else:
        rear_wages.append(0)
        #========================================================================
    rearind=frappe.db.get_all('Broiler Indirect Expenses',filters={'company':layer.company},fields=['title','name'])
    if rearind:
        rear_ind_expanse_item=[]
        rear_ind_itm_cnt=len(rearind)
        for re in rearind:
            rear_ind_expanse_data={}
            acc=[]
            accs=''
            exp=0
            ind_expanse=0
            accsre=frappe.db.sql("""select account from `tabExpense Accounts` where parenttype='Broiler Indirect Expenses' and parent='{0}' """.format(re.name),as_dict=1)
            if accsre:
                for ac in accsre:
                    acc.append(ac.account)
                accs='","'.join(acc)
            exppjt=frappe.db.sql(""" select IFNULL(sum(debit_in_account_currency), 0) as debit,IFNULL(sum(credit_in_account_currency), 0) as credit from `tabGL Entry` where is_cancelled!=1 and account in ("{0}") and posting_date between '{1}' and '{2}' and  project='{3}' """.format(accs,start,end,layer.name),as_dict=1,debug=0)
            if exppjt:
                ind_expanse+=float(exppjt[0].debit)-float(exppjt[0].credit)

            expwpjt=frappe.db.sql(""" select IFNULL(sum(debit_in_account_currency), 0) as debit,IFNULL(sum(credit_in_account_currency), 0) as credit from `tabGL Entry` where is_cancelled!=1 and account in ("{0}") and posting_date between '{1}' and '{2}' and project is NULL """.format(accs,start,end),as_dict=1,debug=0)
            if expwpjt:
                exp=float(expwpjt[0].debit)-float(expwpjt[0].credit)
                if exp != 0:
                    if wageper:
                        ind_expanse+=(float(exp)/100)*float(wageper)                
                    else:
                        batchcount=frappe.db.sql(""" select count(name) as cnt from `tabLayer Batch` where  status='Open' """,as_dict=1,debug=0)
                        if batchcount:
                            ind_expanse+=float(exp)/float(batchcount[0].cnt)

            rear_ind_expanse_tot+=ind_expanse
            col_tot+=ind_expanse
            rear_ind_expanse_data.update({'title':re.title,'amt':ind_expanse})
            rear_ind_expanse_item.append(rear_ind_expanse_data)
        rear_ind_expanse.append(rear_ind_expanse_item)
     
    #----------------------------------
        reat_tot.append(col_tot)
    #------------------------------
    rearing_ind_array=[]
    if rear_ind_itm_cnt:
        for num in range(0,rear_ind_itm_cnt):
            i=0
            row_tot=0
            row_ind=[]
            for indx in rear_ind_expanse:                
                if i==0:
                    row_ind.append(indx[num].get('title'))
                    row_ind.append(indx[num].get('amt'))
                    row_tot+=float(indx[num].get('amt'))
                else:
                    row_ind.append(indx[num].get('amt'))
                    row_tot+=float(indx[num].get('amt'))
                i+=1
            row_ind.append(row_tot)
            rearing_ind_array.append(row_ind)
            
    
    
    reat_tot_tot=float(rear_doc_tot)+float(rear_feed_tot)+float(rear_vaccine_tot)+float(rear_medicine_tot)+float(rear_bio_tot)+float(rear_miscel_tot)+float(rear_other_tot)+float(rear_wages_tot)+float(rear_ind_expanse_tot)
    rear_lbl.append('Total')
    rear_doc.append(rear_doc_tot)
    rear_feed.append(rear_feed_tot)
    rear_vaccine.append(rear_vaccine_tot)
    rear_medicine.append(rear_medicine_tot)
    #rear_bio.append(rear_bio_tot)
    #rear_miscel.append(rear_miscel_tot)
    rear_other.append(rear_other_tot)
    rear_wages.append(rear_wages_tot)
    reat_tot.append(reat_tot_tot)
    rear_graph=[]
    if rear_doc_tot:
        rear_graph.append({"label":"Doc","data":flt(rear_doc_tot,2)})
    if rear_feed_tot:
        rear_graph.append({"label":"Feed","data":flt(rear_feed_tot,2)})
    if rear_vaccine_tot:
        rear_graph.append({"label":"Vaccine","data":flt(rear_vaccine_tot,2)})
    if rear_medicine_tot:
        rear_graph.append({"label":"Medicine","data":flt(rear_medicine_tot,2)})
    if rear_other_tot:
        rear_graph.append({"label":"Other","data":flt(rear_other_tot,2)})
    if rear_wages_tot:
        rear_graph.append({"label":"Wages","data":flt(rear_wages_tot,2)})

#======================================

    rear_html+='<tr class="table-secondary">'
    for lbl in rear_lbl:
        rear_html+='<th scope="col">'+lbl+'</th>'
    rear_html+='</tr>'
    #--------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_doc:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right">'+str(flt(doc,2))+'</td>'
        i+=1

    rear_html+='</tr>'
    #-----------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_feed:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right">'+str(flt(doc,2))+'</td>'
        i+=1

    rear_html+='</tr>'
    #-----------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_vaccine:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right">'+str(flt(doc,2))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_medicine:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right">'+str(flt(doc,2))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    
    #---------------------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_other:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right">'+str(flt(doc,2))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_wages:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right">'+str(flt(doc,2))+'</td>'
        i+=1

    rear_html+='</tr>'
    #-------------------------------------------------
    if len(rearing_ind_array):
        rear_html+='<tr class="table-secondary">'
        i=0
        for lbl in rear_lbl:
            if i==0:
                rear_html+='<th scope="col">Indirect Expense</th>'
            else:
                rear_html+='<td class="text-right"></td>'
            i+=1
        rear_html+='</tr>'
        
        
        for redind in rearing_ind_array:
            rear_html+='<tr >'
            i=0
            for reind in redind:
                if i==0:
                    rear_html+='<th scope="col">'+str(reind)+'</th>'
                else:
                    rear_html+='<td class="text-right">'+str(flt(reind,2))+'</td>'
                i+=1
            rear_html+='</tr>'

    #---------------------------------------------------
    rear_html+='<tr class="table-secondary">'
    i=0
    for doc in reat_tot:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right">'+str(flt(doc,2))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    #Batch Type
    #break_even_period

    

    #------------------------blank ---------------------------
      
    

#=======================================================================
    

    #Layer Budget
    budget=frappe.db.get_value("Broiler Budget",layer.name,['doc','feed','medicine','vaccine','wages','others','production','sales'],as_dict=1)
    doc=0
    feed=0
    medicine=0
    vaccine=0
    wages=0
    others=0
    production=0
    sales=0
    if budget:
       
        doc=budget.doc
        feed=budget.feed
        medicine=budget.medicine
        vaccine=budget.vaccine
        wages=budget.wages
        others=budget.others
        production=budget.production
        sales=budget.sales
    budget_html='' 
    budget_html+='<tr><th>Doc</th><td>'+str(doc)+'</td></tr>'      
    budget_html+='<tr><th>Feed</th><td>'+str(feed)+'</td></tr>'    
    budget_html+='<tr><th>vaccine<td>'+str(vaccine)+'</td></tr>'    
    budget_html+='<tr><th>Medicine</th><td>'+str(medicine)+'</td></tr>'    
    budget_html+='<tr><th>Wages</th><td>'+str(wages)+'</td></tr>'    
    budget_html+='<tr><th>Other</th><td>'+str(others)+'</td></tr>'    
    budget_html+='<tr><th>Production</th><td>'+str(production)+'</td></tr>'    
    budget_html+='<tr><th>Sales</th><td>'+str(sales)+'</td></tr>'

    lay_html=''
    manu_sql=frappe.db.sql(""" select sum(d.transfer_qty) as qty,sum(d.amount) as amount,d.item_code from `tabStock Entry Detail` d left join `tabStock Entry` s 
                on d.parent=s.name where s.stock_entry_type='Manufacture' and d.is_finished_item=1 and 
                s.docstatus=1 and s.manufacturing_type='Broiler Chicken' and s.project='{0}'
                 group by d.item_code""".format(layer.name),as_dict=1,debug=0)
    
    lay_html+='<tr>'
    lay_html+='<th scope="row">Production Qty</th>'
    lay_html+='<td class="text-right">&nbsp;&nbsp;</td>'
    lay_html+='</tr>'
    for man in manu_sql:
        lay_html+='<tr>'
        lay_html+='<th scope="row">'+str(getitem_name(man.item_code))+'</th>'
        lay_html+='<td class="text-right">'+str(man.amount)+'</td>'
        lay_html+='</tr>'

    
    return {'rear':rear_html,'lay':lay_html,'budget':budget_html,'rear_graph':rear_graph}

def getitem_name(item_code):
    return frappe.db.get_value('Item',item_code,'item_name')


import os
#import tempfile


#from frappe.utils.response import download_file
@frappe.whitelist()
def down_report(company,batch,rearing=None,laying=None,budget=None,rearing_gp=None,laying_gp=None,rearing_mor_gp=None,laying_mor_gp=None,rearing_feed_gp=None,laying_feed_gp=None,rearing_weight_gp=None,laying_weight_gp=None,laying_performance_gp=None):
    import json
    import pandas as pd
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook    
    from openpyxl.chart import (
        PieChart, LineChart,
        ProjectedPieChart,
        Reference
    )
    
    from openpyxl.styles import Font
    from openpyxl.chart.axis import DateAxis
    #from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Border, Side

    ft = Font(bold=True)

    rearingary=json.loads(rearing)
    layingary=json.loads(laying)
    budgetary=json.loads(budget)
    gp_rear=json.loads(rearing_gp)
    gp_lay=json.loads(laying_gp)
    gp_rear_mor=json.loads(rearing_mor_gp)
    gp_lay_mor=json.loads(laying_mor_gp)

    gp_rearing_feed=json.loads(rearing_feed_gp),
    gp_laying_feed=json.loads(laying_feed_gp),
    gp_rearing_weight=json.loads(rearing_weight_gp),
    gp_laying_weight=json.loads(laying_weight_gp),
    gp_laying_performance=json.loads(laying_performance_gp)

    rrowlen=len(rearingary)
    rcollen=len(rearingary[0])
    lrowlen=len(layingary)
    lcollen=len(layingary[0])
    browlen=len(budgetary)
    bcollen=len(budgetary[0])
    report=[]
    
    rear_mor=[]
    lbl=['Age','Std. Mortality %','Actual Mortality %']
    #report.append(['Age','Std. Mortality %','Actual Mortality %','Age','Std. Min Feed','Std. Max Feed','Actual Feed','Age','Std. Min Weigt','Std. Max Weight','Actual Weight','Age','Std. Min Eggs','Std. Max Eggs','Actual Eggs'])

    rear_mor.append(lbl)
    j=0
    if len(gp_rear_mor):        
        for rm in gp_rear_mor:
            rear_mor.append([rm.get('age'),rm.get('mortality'),rm.get('act_mortality')])            
            #report.append([rm.get('age'),rm.get('mortality'),rm.get('act_mortality'),rm.get('age'),gp_rearing_feed[0][j].get('v1'),gp_rearing_feed[0][j].get('v2'),gp_rearing_feed[0][j].get('act_feed'),rm.get('age'),gp_rearing_weight[0][j].get('v1'),gp_rearing_weight[0][j].get('v2'),gp_rearing_weight[0][j].get('act_weight'),rm.get('age'),0,0,0])
            
            j+=1
    lay_mor=[]
    lay_mor.append(lbl)
    k=0
    #frappe.msgprint(str(gp_rearing_weight))
    if len(gp_lay_mor):        
        for lm in gp_lay_mor:
            lay_mor.append([lm.get('age'),lm.get('mortality'),lm.get('act_mortality')])
            #report.append([lm.get('age'),lm.get('mortality'),lm.get('act_mortality'),lm.get('age'),gp_laying_feed[0][k].get('v1'),gp_laying_feed[0][k].get('v2'),gp_laying_feed[0][k].get('act_feed'),lm.get('age'),gp_laying_weight[0][k].get('v1'),gp_laying_weight[0][k].get('v2'),gp_laying_weight[0][k].get('act_weight'),lm.get('age'),gp_laying_performance[k].get('v1'),gp_laying_performance[k].get('v2'),gp_laying_performance[k].get('act_eggs')])
            k+=1
            
    rear_exp=[]
    lbl=['','Expenses','Totals']
    rear_exp.append(lbl)
    if len(gp_rear):        
        for rm in gp_rear:
            rear_exp.append(['',rm.get('label'),rm.get('data')])            
          
    lay_exp=[]
    lay_exp.append(lbl)
    if len(gp_lay):        
        for lm in gp_lay:
            lay_exp.append(['',lm.get('label'),lm.get('data')])

    gprearing_feed=[]
    gprearing_feed.append(['Age','Std. Min Feed','Std. Max Feed','Actual Feed'])
    if len(gp_rearing_feed):        
        for lm in gp_rearing_feed[0]:
            gprearing_feed.append([lm.get('age'),float(lm.get('v1')),float(lm.get('v2')),float(lm.get('act_feed'))])

    gplaying_feed=[]
    gplaying_feed.append(['Age','Std. Min Feed','Std. Max Feed','Actual Feed'])
    if len(gp_laying_feed):        
        for lm in gp_laying_feed[0]:
            gplaying_feed.append([lm.get('age'),float(lm.get('v1')),float(lm.get('v2')),float(lm.get('act_feed'))])
    
    gprearing_weight=[]
    gprearing_weight.append(['Age','Std. Min Weigt','Std. Max Weight','Actual Weight'])
    if len(gp_rearing_weight):        
        for lm in gp_rearing_weight[0]:
            gprearing_weight.append([lm.get('age'),float(lm.get('v1')),float(lm.get('v2')),float(lm.get('act_weight'))])

    gplaying_weight=[]
    gplaying_weight.append(['Age','Std. Min Weigt','Std. Max Weight','Actual Weight'])
    if len(gp_laying_weight):        
        for lm in gp_laying_weight[0]:
            gplaying_weight.append([lm.get('age'),float(lm.get('v1')),float(lm.get('v2')),float(lm.get('act_weight'))])

    gplaying_performance=[]
    gplaying_performance.append(['Age','Std. Min Eggs','Std. Max Eggs','Actual Eggs'])
    if len(gp_laying_performance):        
        for lm in gp_laying_performance:
            gplaying_performance.append([lm.get('age'),float(lm.get('v1')),float(lm.get('v2')),float(lm.get('act_eggs'))])

    wb = Workbook()
    ws = wb.active
    ws.title = "Rearing"
    ws2 = wb.create_sheet("Production")
    ws1 = wb.create_sheet("Budget")   
    ws5 = wb.create_sheet("Rear. Mor. GPH")
    ws6 = wb.create_sheet("Lay. Mor. GPH")
    #rep = wb.create_sheet("Report")
    ws7 = wb.create_sheet("Rear. Feed. GPH")
    ws8 = wb.create_sheet("Lay. Feed. GPH")
    ws9 = wb.create_sheet("Rear. Weight. GPH")
    ws10 = wb.create_sheet("Lay. Weight. GPH")
    ws11 = wb.create_sheet("Performance. GPH")
    

    for row in rearingary:
       ws.append(row)

    for row in budgetary:
       ws1.append(row)

    for row in layingary:
       ws2.append(row)
    
    ws.append([''])
    for row in rear_exp:
        ws.append(row)
       
    ws2.append([''])
    for row in lay_exp:
        ws2.append(row)
        
    for row in rear_mor:
        ws5.append(row)

    for row in lay_mor:
        ws6.append(row)

    for row in gprearing_feed:
        ws7.append(row)

    for row in gplaying_feed:
        ws8.append(row)
    
    for row in gprearing_weight:
        ws9.append(row)

    for row in gplaying_weight:
        ws10.append(row)

    for row in gplaying_performance:
        ws11.append(row)


    #for row in report:
     #   rep.append(row)
    
    yellow = "00D5D7D9"
    black="00000000"
    thin = Side(border_style="thin", color=black)
    double = Side(border_style="double", color=black)
    thick = Side(border_style="thick", color=black)
    #=====================================================
    """"
    for row in rep['A1:O1']:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    for row in rep['C1:C101']:
        for cell in row:
            cell.border = Border(right=thick)
    for row in rep['G1:G101']:
        for cell in row:
            cell.border = Border(right=thick)
    for row in rep['K1:K101']:
        for cell in row:
            cell.border = Border(right=thick)
            """
    
    #===========================================================
    rlbl=getColumnName(rcollen)
    rhd="A1:"+str(rlbl)+str(1)
    
    for row in ws[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:"+str(rlbl)+str(rrowlen)
    for row in ws[rhd]:
        rowhed=0
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if "Total" in str(cell.value):
                rowhed=1
            if rowhed==1:
                cell.font = ft
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:A"+str(rrowlen)
    for row in ws[rhd]:
        for cell in row:
            #cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    #---------------------------------------------------------------------------------
    rlbl=getColumnName(bcollen)
    rhd="A1:"+str(rlbl)+str(1)
    
    for row in ws1[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:"+str(rlbl)+str(browlen)
    for row in ws1[rhd]:
        rowhed=0
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if "Total" in str(cell.value):
                rowhed=1
            if rowhed==1:
                cell.font = ft
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:A"+str(browlen)
    for row in ws1[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    #----------------------------------------------------------------------------------------------
    rlbl=getColumnName(lcollen)
    rhd="A1:"+str(rlbl)+str(1)
    for row in ws2[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:"+str(rlbl)+str(lrowlen)
    for row in ws2[rhd]:
        rowhed=0
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if "Total" in str(cell.value) or 'Net Profit' in str(cell.value) or 'Operational Cost' in str(cell.value):
                rowhed=1
            if rowhed==1:
                cell.font = ft
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    
    rhd="A1:A"+str(lrowlen)
    for row in ws2[rhd]:
        for cell in row:
            #cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    if len(gp_rear):
        maxr=len(gp_rear)+1
        rhd="B"+str(rrowlen+2)
        pie = PieChart()
        #labels = Reference(ws3, min_col=2, min_row=2, max_row=maxr)
        labels = "'Rearing'!$B$"+str(rrowlen+3)+":$B$"+str(rrowlen+1+maxr)
        #frappe.msgprint(str(labels))
        #data = Reference(ws3, min_col=3, min_row=1, max_row=maxr)
        data = "'Rearing'!$C$"+str(rrowlen+3)+":$C$"+str(rrowlen+1+maxr)
        #frappe.msgprint(str(data))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Rearing"
        ws.add_chart(pie, rhd)
        
    if len(gp_lay):
        maxr=len(gp_lay)+1
        rhd="B"+str(lrowlen+2)
        pie = PieChart()        
        labels = "'Production'!$B$"+str(lrowlen+3)+":$B$"+str(lrowlen+1+maxr)  
        data = "'Production'!$C$"+str(lrowlen+3)+":$C$"+str(lrowlen+1+maxr)        
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Laying"
        ws2.add_chart(pie, rhd)
    from copy import deepcopy
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import RichTextProperties

    if len(gp_rear_mor):
        
        maxr=len(gp_rear_mor)+1
        c2 = LineChart()
        c2.title = "Rearing Mortality"
        c2.style = 13
        c2.x_axis.title = 'Weeks'
        c2.y_axis.title = 'Mortality %'
        c2.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c2.x_axis.delete = False
        c2.y_axis.delete = False
        c2.height = 10 # default is 7.5
        c2.width = 25 # default is 15
        data = Reference(ws5, min_col=2, min_row=1, max_col=3, max_row=maxr)
        #frappe.msgprint(str(data)) #= 'Rear. Mor. GPH'!$B$1:$C$18
        c2.add_data(data, titles_from_data=True)
        dates = Reference(ws5, min_col=1, min_row=2, max_row=maxr)
        #frappe.msgprint(str(dates)) #='Rear. Mor. GPH'!$A$2:$A$18
        c2.set_categories(dates)
        s2 = c2.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c2.series[1]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs
        ws5.add_chart(c2, "A1")

        #=========================
    if len(gp_lay_mor):
        maxr=len(gp_lay_mor)+1
        rhd="A1"
        
        c1 = LineChart()
        c1.title = "Laying Mortality"
        c1.style = 13
        c1.x_axis.title = 'Weeks'
        c1.y_axis.title = 'Mortality %'
        c1.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c1.x_axis.delete = False
        c1.y_axis.delete = False
        c1.height = 20 # default is 7.5
        c1.width = 60 # default is 15
        data = Reference(ws6, min_col=2, min_row=1, max_col=3, max_row=maxr)
        c1.add_data(data, titles_from_data=True)
        dates = Reference(ws6, min_col=1, min_row=2, max_row=maxr)
        c1.set_categories(dates)
        s2 = c1.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c1.series[1]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs        
        ws6.add_chart(c1, rhd)
        
    #----------------------------------------------------

    if len(gprearing_feed):
        
        maxr=len(gprearing_feed)
        c3 = LineChart()
        c3.title = "Rearing Feed"
        c3.style = 13
        c3.x_axis.title = 'Weeks'
        c3.y_axis.title = 'Feed'
        c3.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c3.x_axis.delete = False
        c3.y_axis.delete = False
        c3.height = 10 # default is 7.5
        c3.width = 25 # default is 15
        data = Reference(ws7, min_col=2, min_row=1, max_col=4, max_row=maxr)            
        c3.add_data(data, titles_from_data=True)
        dates = Reference(ws7, min_col=1, min_row=2, max_row=maxr)        
        c3.set_categories(dates)
        s2 = c3.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c3.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c3.series[2]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs        
       

        ws7.add_chart(c3, "A1")

        #=========================
    if len(gplaying_feed):
        maxr=len(gplaying_feed)
        rhd="A1"
        
        c4 = LineChart()
        c4.title = "Laying Feed"
        c4.style = 13
        c4.x_axis.title = 'Weeks'
        c4.y_axis.title = 'Feed'
        c4.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c4.x_axis.delete = False
        c4.y_axis.delete = False
        c4.height = 20 # default is 7.5
        c4.width = 60 # default is 15
        data = Reference(ws8, min_col=2, min_row=1, max_col=4, max_row=maxr)
        c4.add_data(data, titles_from_data=True)
        dates = Reference(ws8, min_col=1, min_row=2, max_row=maxr)
        c4.set_categories(dates)
        s2 = c4.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c4.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c4.series[2]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs 

        ws8.add_chart(c4, rhd)

    #----------------------------------------------------

    if len(gprearing_weight):
        
        maxr=len(gprearing_weight)
        c3 = LineChart()
        c3.title = "Rearing Weight"
        c3.style = 13
        c3.x_axis.title = 'Weeks'
        c3.y_axis.title = 'Weight'
        c3.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c3.x_axis.delete = False
        c3.y_axis.delete = False
        c3.height = 10 # default is 7.5
        c3.width = 25 # default is 15
        data = Reference(ws9, min_col=2, min_row=1, max_col=4, max_row=maxr)            
        c3.add_data(data, titles_from_data=True)
        dates = Reference(ws9, min_col=1, min_row=2, max_row=maxr)        
        c3.set_categories(dates)
        s2 = c3.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c3.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c3.series[2]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs        
       

        ws9.add_chart(c3, "A1")

        #=========================
    if len(gplaying_weight):
        maxr=len(gplaying_weight)
        rhd="A1"
        
        c4 = LineChart()
        c4.title = "Laying Weight"
        c4.style = 13
        c4.x_axis.title = 'Weeks'
        c4.y_axis.title = 'Weight'
        c4.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c4.x_axis.delete = False
        c4.y_axis.delete = False
        c4.height = 20 # default is 7.5
        c4.width = 60 # default is 15
        data = Reference(ws10, min_col=2, min_row=1, max_col=4, max_row=maxr)
        c4.add_data(data, titles_from_data=True)
        dates = Reference(ws10, min_col=1, min_row=2, max_row=maxr)
        c4.set_categories(dates)
        s2 = c4.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c4.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c4.series[2]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs 

        ws10.add_chart(c4, rhd)

    if len(gplaying_performance):
        maxr=len(gplaying_performance)
        rhd="A1"
        
        c4 = LineChart()
        c4.title = "Laying Performance"
        c4.style = 13
        c4.x_axis.title = 'Weeks'
        c4.y_axis.title = 'Eggs'
        c4.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c4.x_axis.delete = False
        c4.y_axis.delete = False
        c4.height = 20 # default is 7.5
        c4.width = 60 # default is 15
        data = Reference(ws11, min_col=2, min_row=1, max_col=4, max_row=maxr)
        c4.add_data(data, titles_from_data=True)
        dates = Reference(ws11, min_col=1, min_row=2, max_row=maxr)
        c4.set_categories(dates)
        s2 = c4.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c4.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c4.series[2]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs 2490ef

        ws11.add_chart(c4, rhd)

    file_name = 'poultry_dash.xlsx'    
    temp_file=os.path.join(frappe.utils.get_bench_path(), "logs", file_name)
    wb.save(temp_file)
    return temp_file

def getColumnName(n):
 
    # initialize output string as empty
    result = ''
 
    while n > 0:
 
        # find the index of the next letter and concatenate the letter
        # to the solution
 
        # here index 0 corresponds to 'A', and 25 corresponds to 'Z'
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26
 
    return result[::-1]


@frappe.whitelist()
def down_file(file=None):
    from frappe.utils.file_manager import download_file
    file_name = os.path.basename(file)
    with open(file, "rb") as fileobj:
        filedata = fileobj.read()
    frappe.response['content_type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    frappe.response['content_disposition'] = 'attachment; filename="{0}"'.format(file_name)
    frappe.local.response.filename = file_name
    frappe.local.response.filecontent = filedata
    frappe.local.response.type = "download"

@frappe.whitelist()
def get_rear_mor_graph(batch,period):
    bth=frappe.db.get_value('Layer Batch',batch,['strain','doc_placed_date','flock_transfer_date','doc_placed'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.doc_placed_date)
    actmort=frappe.db.get_all('Rearing Period Performance',filters={'parent':strain},fields=['age','mortality'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,sum(total) as mort FROM  `tabLayer Mortality` 
    WHERE parentfield='rearing_daily_mortality' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
   
    date2=date
    for ly in actmort:
    
        date2=add_days(date2,7)
        wkd=date_diff(date2,date)//7
        mortp=0
        for ac in actqry:
            if ac.wk==wkd:
                mortp=(float(ac.mort)*100)/doc_placed
        ly.update({'act_mortality':mortp})

    return {'ideal':actmort}

@frappe.whitelist()
def get_lay_mor_graph(batch,period):
    bth=frappe.db.get_value('Layer Batch',batch,['strain','doc_placed_date','flock_transfer_date','doc_placed'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.flock_transfer_date)
    actmort=frappe.db.get_all('Laying Period Performance',filters={'parent':strain},fields=['age','mortality'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,sum(total) as mort FROM  `tabLayer Mortality` 
    WHERE parentfield='laying_mortality' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
   
    mortidel=[]
    date2=date
    maxweek=0
    if actqry:
        maxweek=actqry[-1].wk

    for ly in actmort:
        date2=add_days(date2,7)
        wkd=date_diff(date2,date)//7
        #if maxweek < wkd:
         #  break
        
        mortp=0
        for ac in actqry:
            if ac.wk==wkd:
                mortp=(float(ac.mort)*100)/doc_placed
                               
        ly.update({'act_mortality':mortp})
        mortidel.append(ly)
    
    return {'ideal':mortidel}

@frappe.whitelist()
def get_lay_feed_graph(batch,period):
    bth=frappe.db.get_value('Layer Batch',batch,['strain','doc_placed_date','flock_transfer_date','doc_placed'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.flock_transfer_date)
    stdfeed=frappe.db.get_all('Laying Period Performance',filters={'parent':strain},fields=['age','TRIM(SUBSTRING_INDEX(feed_intake,"–",1)) as v1','TRIM(SUBSTRING_INDEX(feed_intake,"–",-1)) as v2'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(qty*conversion_factor),0) as qty FROM  `tabLayer Feed` 
    WHERE parentfield='laying_feed' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
    rear_mortality=0
    rear_mortality_sql=frappe.db.sql(""" SELECT IFNULL(sum(total),0) as rear_mort FROM  `tabLayer Mortality` 
    WHERE parentfield='rearing_daily_mortality' and parent='{0}'  """.format(batch),as_dict=1,debug=0)
    if rear_mortality_sql:
        rear_mortality+=rear_mortality_sql[0].rear_mort

    mortality=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(total),0) as mort FROM  `tabLayer Mortality` 
    WHERE parentfield='laying_mortality' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)

    mortidel=[]
    date2=date
    maxweek=0
    if actqry:
        maxweek=actqry[-1].wk

    for ly in stdfeed:
        date2=add_days(date2,7)
        wkd=date_diff(date2,date)//7
        
        #if maxweek < wkd:
         #  break
        mort=0
        for mo in mortality:
            if wkd<=mo.wk:
                mort+=float(mo.mort)
            

        feed=0
        for ac in actqry:
            if ac.wk==wkd:
                livechick=float(doc_placed)-float(mort)+float(rear_mortality)
                feed=float(ac.qty*1000000)/(float(livechick)*7)
                              
        ly.update({'act_feed':feed})
        mortidel.append(ly)
    
    return {'ideal':mortidel}

@frappe.whitelist()
def get_lay_weight_graph(batch,period):
    bth=frappe.db.get_value('Layer Batch',batch,['strain','doc_placed_date','flock_transfer_date','doc_placed'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.flock_transfer_date)
    stdweight=frappe.db.get_all('Laying Period Performance',filters={'parent':strain},fields=['age','TRIM(SUBSTRING_INDEX(body_weight,"–",1)) as v1','TRIM(SUBSTRING_INDEX(body_weight,"–",-1)) as v2'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(weight),0) as weight FROM  `tabLayer Weight` 
    WHERE parentfield='laying_weight' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
    
    mortidel=[]
    date2=date
    maxweek=0
    if actqry:
        maxweek=actqry[-1].wk

    for ly in stdweight:
        date2=add_days(date2,7)
        wkd=date_diff(date2,date)//7
        #if maxweek < wkd:
         #  break
        
        mortp=0
        for ac in actqry:
            if ac.wk==wkd:
                mortp=ac.weight
                               
        ly.update({'act_weight':mortp})
        mortidel.append(ly)
    
    return {'ideal':mortidel}

@frappe.whitelist()
def get_rear_feed_graph(batch,period):
    bth=frappe.db.get_value('Layer Batch',batch,['strain','doc_placed_date','flock_transfer_date','doc_placed'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.doc_placed_date)
    actmort=frappe.db.get_all('Rearing Period Performance',filters={'parent':strain},fields=['age','TRIM(SUBSTRING_INDEX(feed_intake,"–",1)) as v1','TRIM(SUBSTRING_INDEX(feed_intake,"–",-1)) as v2'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(qty*conversion_factor),0) as qty FROM  `tabLayer Feed` 
    WHERE parentfield='rearing_feed' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
   
    mortality=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(total),0) as mort FROM  `tabLayer Mortality` 
    WHERE parentfield='rearing_daily_mortality' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
    date2=date
    for ly in actmort:
    
        date2=add_days(date2,7)
        wkd=date_diff(date2,date)//7
        
        mort=0
        for mo in mortality:
            if wkd<=mo.wk:
                mort+=float(mo.mort)
            
        feed=0
        livechick=0
        for ac in actqry:
            if ac.wk==wkd:
                livechick=float(doc_placed)-float(mort)
                feed=float(ac.qty*1000000)/(float(livechick)*7)
        ly.update({'act_feed':feed})
    return {'ideal':actmort}

@frappe.whitelist()
def get_rear_weight_graph(batch,period):
    bth=frappe.db.get_value('Layer Batch',batch,['strain','doc_placed_date','flock_transfer_date','doc_placed'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.doc_placed_date)
    actmort=frappe.db.get_all('Rearing Period Performance',filters={'parent':strain},fields=['age','TRIM(SUBSTRING_INDEX(body_weight,"–",1)) as v1','TRIM(SUBSTRING_INDEX(body_weight,"–",-1)) as v2'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(weight),0) as weight FROM  `tabLayer Weight` 
    WHERE parentfield='rearing_weight' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
   
    date2=date
    for ly in actmort:
    
        date2=add_days(date2,7)
        wkd=date_diff(date2,date)//7
        mortp=0
        for ac in actqry:
            if ac.wk==wkd:
                mortp=ac.weight
        ly.update({'act_weight':mortp})

    return {'ideal':actmort}
    
@frappe.whitelist()
def get_lay_performance(batch,period):
    bth=frappe.db.get_value('Layer Batch',batch,['strain','doc_placed_date','flock_transfer_date','doc_placed'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.flock_transfer_date)
    stdeggs=frappe.db.get_all('Laying Period Performance',filters={'parent':strain},fields=['age','TRIM(SUBSTRING_INDEX(hen_housed_eggs,"–",1)) as v1','TRIM(SUBSTRING_INDEX(hen_housed_eggs,"–",-1)) as v2'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(qty*conversion_factor),0) as qty FROM  `tabEgg Production` 
    WHERE  parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
    rear_mortality=0
    rear_mortality_sql=frappe.db.sql(""" SELECT IFNULL(sum(total),0) as rear_mort FROM  `tabLayer Mortality` 
    WHERE parentfield='rearing_daily_mortality' and parent='{0}'  """.format(batch),as_dict=1,debug=0)
    if rear_mortality_sql:
        rear_mortality+=rear_mortality_sql[0].rear_mort

    mortality=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') DIV 7 as wk,IFNULL(sum(total),0) as mort FROM  `tabLayer Mortality` 
    WHERE parentfield='laying_mortality' and parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)

    mortidel=[]
    date2=date
    maxweek=0
    if actqry:
        maxweek=actqry[-1].wk

    for ly in stdeggs:
        date2=add_days(date2,7)
        wkd=date_diff(date2,date)//7
        
        #if maxweek < wkd:
         #  break
        mort=0
        for mo in mortality:
            if wkd<=mo.wk:
                mort+=float(mo.mort)
            

        eggs=0
        for ac in actqry:
            if ac.wk==wkd:
                livechick=float(doc_placed)-float(mort)+float(rear_mortality)
                eggs=float(ac.qty)/(float(livechick)*7)
                              
        ly.update({'act_eggs':eggs})
        mortidel.append(ly)
    
    return {'ideal':mortidel}