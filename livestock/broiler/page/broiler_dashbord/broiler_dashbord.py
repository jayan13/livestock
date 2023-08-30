import frappe
import erpnext
from frappe.utils import getdate,add_days,get_first_day,get_last_day,nowdate,flt,date_diff
from erpnext.stock.get_item_details import (
	get_conversion_factor,
	get_default_cost_center,
)
from erpnext.stock.utils import (get_incoming_rate)
from erpnext.stock.doctype.item.item import get_item_defaults 

@frappe.whitelist()
def get_company_list():
    data = {}
    data["companys"] = frappe.get_list("Company", fields=['name'],limit_page_length=0, order_by="name",debug=0)
    return data

@frappe.whitelist()
def get_batch_list(company=None,status=None):
    data = {}
    data["batchs"] = frappe.get_all("Broiler Batch",filters={'docstatus':['!=','2'],'company':company,'status':status},fields=['name'],limit_page_length=0, order_by="start_date desc",debug=0)
    return data

@frappe.whitelist()
def get_sale_item_list():
    itemgp=['CHICKEN PRODUCTS - ACACIA','CHICKEN PRODUCTS - AL FAKHER','CHICKEN PRODUCTS - AUH']
    return frappe.db.get_all('Item',filters={'item_group':['in',itemgp]},fields=['item_code'],pluck='item_code')

def get_item_price(company,item,warehouse,date,qty,uom):
    conversion_factor = get_conversion_factor(item,uom).get("conversion_factor") or 1
    base_rate= get_incoming_rate({
						"item_code": item,
						"warehouse": warehouse,
						"posting_date": date,
						"posting_time": '23:59:59',
						"qty": -1 * qty,
                        'company':company
					}) 
    return base_rate*conversion_factor

def get_item_baseprice(company,item,warehouse,date,qty):
    #item_account_details = get_item_defaults(item, company)
    #stock_uom = item_account_details.stock_uom
    #conversion_factor = get_conversion_factor(item,stock_uom).get("conversion_factor") or 1
    base_rate= get_incoming_rate({
						"item_code": item,
						"warehouse": warehouse,
						"posting_date": date,
						"posting_time": '23:59:59',
						"qty": -1 * qty,
                        'company':company
					}) 
    return base_rate

@frappe.whitelist()
def get_report(company,batch):
    layer=frappe.get_doc('Broiler Batch',batch)
    default_currency=frappe.get_value('Company',layer.company,'default_currency')
    
    pjt=frappe.db.get_value('Project',layer.project,['expected_start_date','expected_end_date'], as_dict=1)
   
    end_date=layer.end_date
    if pjt:
        if not layer.end_date:
            end_date=getdate(pjt.expected_end_date) or nowdate()

    rear_perid=[]
    
    doc_placed_date=getdate(layer.start_date)
    rear_start_date=doc_placed_date
    rear_end_date=end_date
    rstart=rear_start_date
    rend=''
    dept=[]
    depts=frappe.get_doc('Broiler Wage Expense Departments',layer.company)
    if depts:
        for dep in depts.department:
            dept.append(dep.name)
    #lay_days=date_diff(lay_end_date,lay_start_date)+1
    
    #----- rearing 
    rear_lbl=['Expense']
    rear_doc=['Doc']
    rear_feed=['Feed']
    rear_vaccine=['Vaccine']
    rear_medicine=['Medicine-Disinfectants']
    #rear_bio=['Biosecurity Requirements']
    #rear_miscel=['Miscellaneous Production Req.']
    rear_other=['Others']
    rear_direct=['Direct Material Cost']
    rear_tot_direct=['Total Direct Exp (Direct Material Cost + Wages)']
    rear_wages=['Wages']
    reat_tot=['Total (Direct + Indirect Expense)']
    rear_doc_tot=0
    rear_feed_tot=0
    rear_vaccine_tot=0
    rear_medicine_tot=0
    rear_bio_tot=0
    rear_miscel_tot=0
    rear_other_tot=0
    rear_wages_tot=0
    reat_tot_tot=0
    rear_ind_expanse=[]
    rear_ind_expanse_tot=0
    rear_ind_itm_cnt=0
    
    
    rear_html=''

    vaccine_list=[]
    itmgroup=frappe.db.get_all('Vaccine Item Group',fields=['item_group'],pluck='item_group')
    for gp in itmgroup:
        itms=frappe.db.get_all('Item',filters={'item_group':gp},fields=['item_code'],pluck='item_code')
        for it in itms:
            vaccine_list.append(it)
    
#=========================================================================
    
    
    date_lbl=rear_start_date.strftime("%d/%m/%y")+'-'+rear_end_date.strftime("%d/%m/%y")
    rear_lbl.append(date_lbl)
    col_tot=0
    direct_material_cost=0
        #---------------------------------
    base_row_material=''
    finished_product=''
    raw_warehouse=''
    feed_warehouse=''
    cull=''
    if layer.broiler_shed:
        broiler_shed=frappe.db.get_value('Broiler Shed',layer.broiler_shed,['base_row_material','row_material_target_warehouse','product_target_warehouse','product','feed_warehouse','cull'],as_dict=1)
        if broiler_shed:
            base_row_material=broiler_shed.base_row_material
            finished_product=broiler_shed.product
            raw_warehouse=broiler_shed.row_material_target_warehouse
            feed_warehouse=broiler_shed.feed_warehouse
            cull=broiler_shed.cull

    fromdate=add_days(getdate(rear_start_date),-15)
    
       
       
    #----------------------------------
    manuf=frappe.db.sql(""" select IFNULL(sum(d.amount), 0) as amount,IFNULL(sum(d.qty), 0) as qty,d.item_code from `tabStock Entry Detail` d left join `tabStock Entry` s 
                on d.parent=s.name where s.stock_entry_type='Manufacture' and s.docstatus=1 and s.project='{0}' 
                 group by d.item_code """.format(layer.name),as_dict=1,debug=0)
    amt=0
    price=0
    if base_row_material:
        if layer.number_received:
            qty=0
            for man in manuf:
                #if man.item_code in [base_row_material,cull]:
                if man.item_code==base_row_material:
                    qty+=man.qty
                    amt+=man.amount

            dqty=layer.number_received-qty
            if dqty:
                price=get_item_price(company,base_row_material,raw_warehouse,rear_start_date,dqty,'Nos')
                amt+=price*dqty

            rear_doc.append(amt)
            rear_doc_tot=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_doc.append('0')
    else:
        rear_doc.append('0')

    feeditem=[]
    if layer.feed:
        for itm in layer.feed:
            if itm.starter_item:
                feeditem.append(itm.starter_item)
            if itm.finisher_item:
                feeditem.append(itm.finisher_item)
        feeditem=frappe.utils.unique(feeditem)
    
    meditem=[]
    if layer.medicine:                
        for itm in layer.medicine:
            if itm.item:
                meditem.append(itm.item)
        meditem=frappe.utils.unique(meditem)

    vaccitem=[]
    if layer.vaccine:                
        for itm in layer.vaccine:
            if itm.item:
                vaccitem.append(itm.item)
        vaccitem=frappe.utils.unique(vaccitem)

    otheritem=[]
    if layer.used_items:                
        for itm in layer.used_items:
            if itm.item_code:
                otheritem.append(itm.item_code)
        otheritem=frappe.utils.unique(otheritem)
    
    if not layer.item_processed or layer.item_processed=='0':
        
        process={}
        feeds=[]
        processed_sql=frappe.db.get_all('Broiler Transfer Consumable',filters={'processed':1,'batch':layer.name},fields=['sum(used_quantity) as qty','materal'],group_by='materal') 
        if processed_sql:
            for proc in processed_sql:
                process.update({proc.materal:proc.qty})
                
        
        if layer.feed:
            feedsqls=frappe.db.sql(""" select sum(starter_qty) as qty,starter_item as item,starter_uom as uom,max(`date`) as last_date from `tabFeed` where starter_item is not null and starter_item!='' and parent='{0}' group by CONCAT(starter_item,'-',starter_uom)""".format(layer.name),as_dict=1,debug=0)
            feedsqlf=frappe.db.sql(""" select sum(finisher_qty) as qty,finisher_item as item,finisher_uom as uom,max(`date`) as last_date from `tabFeed` where finisher_item is not null and finisher_item!='' and parent='{0}' group by CONCAT(finisher_item,'-',finisher_uom)""".format(layer.name),as_dict=1,debug=0)
            
            if feedsqls:
                for feed in feedsqls:
                    qty=0
                    price=0
                    data=['0','0','']
                    price=get_item_price(company,feed.item,feed_warehouse,feed.last_date,qty,feed.uom)
                    data[0]=feed.qty
                    data[1]=price
                    data[2]=feed.item
                    feeds.append(data)

            if feedsqlf:
                for feed in feedsqlf:
                    qty=0
                    price=0
                    data=['0','0','']
                    #conversion_factor = get_conversion_factor(feed.item,feed.uom).get("conversion_factor") or 1
                    #qty=feed.qty*conversion_factor
                    price=get_item_price(company,feed.item,feed_warehouse,feed.last_date,qty,feed.uom)
                    data[0]=feed.qty
                    data[1]=price
                    data[2]=feed.item
                    feeds.append(data)

            #frappe.msgprint(str(feeds))
            amt=0
            if feeds:
                for f in feeds:                    
                    qty=0
                    if process.get(f[2]):
                        qty=float(f[0])-float(process.get(f[2]))
                        amt+=qty*float(f[1])
                    else:
                        qty=float(f[0])
                        amt+=qty*float(f[1])

            if manuf:
                for man in manuf:
                    if man.item_code in feeditem:
                        amt+=man.amount
                        
            #frappe.throw(str(feeds))
            rear_feed.append(amt)
            rear_feed_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_feed.append(0)

                #---------------------------------------         
        if layer.medicine:
            medsql=frappe.db.sql(""" select sum(qty) as qty,item,uom as uom,max(`date`) as last_date from `tabMedicine` where item is not null and item!='' and parent='{0}' group by CONCAT(item,'-',uom)""".format(layer.name),as_dict=1,debug=0)
            meds=[]
            if medsql:
                for feed in medsql:
                    qty=0
                    price=0
                    data=['0','0','']
                    price=get_item_price(company,feed.item,base_row_material,feed.last_date,qty,feed.uom)
                    data[0]=feed.qty
                    data[1]=price
                    data[2]=feed.item
                    meds.append(data)

            amt=0
            if meds:
                for f in meds:                    
                    qty=0
                    if process.get(f[2]):
                        qty=float(f[0])-float(process.get(f[2]))
                        amt+=qty*float(f[1])
                    else:
                        qty=float(f[0])
                        amt+=qty*float(f[1])

            if manuf:
                for man in manuf:
                    if man.item_code in meditem:
                        amt+=man.amount

            rear_medicine.append(amt)
            rear_medicine_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_medicine.append(0)

        if layer.vaccine:
            vaccsql=frappe.db.sql(""" select sum(qty) as qty,item,uom as uom,max(`date`) as last_date from `tabVaccine` where item is not null and item!='' and parent='{0}' group by CONCAT(item,'-',uom)""".format(layer.name),as_dict=1,debug=0)
            vccs=[]
            if vaccsql:
                for feed in vaccsql:
                    qty=0
                    price=0
                    data=['0','0','']
                    price=get_item_price(company,feed.item,base_row_material,feed.last_date,qty,feed.uom)
                    data[0]=feed.qty
                    data[1]=price
                    data[2]=feed.item
                    vccs.append(data)

            amt=0
            if vccs:
                for f in vccs:                    
                    qty=0
                    if process.get(f[2]):
                        qty=float(f[0])-float(process.get(f[2]))
                        amt+=qty*float(f[1])
                    else:
                        qty=float(f[0])
                        amt+=qty*float(f[1])

            if manuf:
                for man in manuf:
                    if man.item_code in vaccitem:
                        amt+=man.amount
                    
            rear_vaccine.append(amt)
            rear_vaccine_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_vaccine.append(0)
                    
                #------------------------------------------
        if layer.used_items:
            othsql=frappe.db.sql(""" select sum(qty) as qty,item_code as item,uom as uom,max(`date`) as last_date from `tabBroiler Items` where item_code is not null and item_code!='' and parent='{0}' group by CONCAT(item_code,'-',uom)""".format(layer.name),as_dict=1,debug=0)
            othss=[]
            if othsql:
                for feed in othsql:
                    qty=0
                    price=0
                    data=['0','0','']
                    price=get_item_price(company,feed.item,base_row_material,feed.last_date,qty,feed.uom)
                    data[0]=feed.qty
                    data[1]=price
                    data[2]=feed.item
                    othss.append(data)

            amt=0
            if othss:
                for f in othss:                    
                    qty=0
                    if process.get(f[2]):
                        qty=float(f[0])-float(process.get(f[2]))
                        amt+=qty*float(f[1])
                    else:
                        qty=float(f[0])
                        amt+=qty*float(f[1])

            if manuf:
                for man in manuf:
                    if man.item_code in otheritem:
                        amt+=man.amount
            
            rear_other.append(amt)
            rear_other_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_other.append(0)
    else:
        #---------------- all processed item ---------------------------
        amt=0
        if manuf:
            for man in manuf:
                if man.item_code in feeditem:
                    amt+=man.amount
                        
            rear_feed.append(amt)
            rear_feed_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_feed.append(0)    
        
        if manuf:    
            amt=0
            for man in manuf:
                if man.item_code in meditem:
                    amt+=man.amount
                    
            rear_medicine.append(amt)
            rear_medicine_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_medicine.append(0)

        if manuf:
                    
            amt=0
            for man in manuf:
                if man.item_code in vaccitem:
                    amt+=man.amount

            rear_vaccine.append(amt)
            rear_vaccine_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_vaccine.append(0)
                    
                
        if manuf:
            amt=0
            for man in manuf:
                if man.item_code in otheritem:
                    amt+=man.amount

            rear_other.append(amt)
            rear_other_tot+=amt
            col_tot+=amt
            direct_material_cost+=amt
        else:
            rear_other.append(0)

    #-------------------------------
    rear_direct.append(direct_material_cost)
        #-----------------------------------------
    start=rear_start_date
    end=rear_end_date
    s=get_first_day(rear_start_date)
    e=get_last_day(rear_end_date)
    saldy=date_diff(e,s)
    tdy=date_diff(rear_end_date,rear_start_date)
    deptsql=''
    if len(dept):
        dp='","'.join(dept)
        deptsql=' and department in ("'+str(dp)+'") '

    date_range=frappe.db.sql(""" select min(start_date) as start_date,max(end_date) as end_date from `tabSalary Slip` where status in ('Draft','Submitted') and company='{0}' {1} and  end_date between '{2}' and '{3}' """.format(layer.company,deptsql,s,e),as_dict=1,debug=0)    
    if date_range:
        saldy=date_diff(date_range[0].end_date,date_range[0].start_date)+1
        
    salsql=frappe.db.sql(""" select net_pay,start_date,end_date from `tabSalary Slip` where status in ('Draft','Submitted') and company='{0}' {1} and  end_date between '{2}' and '{3}' """.format(layer.company,deptsql,s,e),as_dict=1,debug=0)
    totsal=0
    if salsql:
        for sal in salsql:
            totsal+=sal.net_pay

    salary=0
    daysal=0
    
    if totsal:
        daysal=float(totsal)/float(saldy)
    
    salary_expanse=0
    #salary=float(daysal)*float(tdy)
    wageper=100
    #find each day live chicken and each day salary. each day batch sal = each day live percentage * each day salary
    st_date=rear_start_date
    salary_expanse=0
    totper=0
    ed_date=rear_end_date
    if getdate(rear_end_date) > getdate(nowdate()):
        ed_date=getdate(nowdate())
        tdy=date_diff(ed_date,rear_start_date)

    while getdate(st_date) < getdate(add_days(ed_date,1)):
        
        tot_live=0
        batch_live=0
        eachday_live=frappe.db.sql("""select b.doc_placed-IFNULL(sum(m.total), 0) as live,b.name from `tabBroiler Batch` b 
        left join `tabMortality` m on b.name=m.parent and m.date <= '{1}' where
            b.company='{0}' and '{1}' between b.receiving_date and b.end_date  group by b.name""".format(layer.company,st_date,layer.name),as_dict=1,debug=0)
        for ech in eachday_live:
            if layer.name==ech.name:
                batch_live+=float(ech.live)
            else:
                tot_live+=float(ech.live)

        salper=(float(batch_live)*100)/float(tot_live)
        totper+=float(salper)
        if totsal:
            daysal=float(totsal)/float(tdy)
            salary_expanse+=(float(daysal)/100)*float(salper)
        st_date=add_days(st_date,1)
        
    wageper=float(totper)/float(tdy)    
    
    if daysal:
        rear_wages.append(salary_expanse)
        rear_wages_tot+=salary_expanse
        col_tot+=salary_expanse
    else:
        rear_wages.append(0)

        #========================================================================

    #-------------------------------
    rear_tot_direct.append((direct_material_cost)+float(rear_wages_tot))
    #-------------------------------
    rearind=frappe.db.get_all('Broiler Indirect Expenses',filters={'company':layer.company},fields=['title','name'])
    if rearind:
        rear_ind_expanse_item=[]
        rear_ind_itm_cnt=len(rearind)
        for re in rearind:
            rear_ind_expanse_data={}
            acc=[]
            accs=''
            exp=0
            ind_expanse=0
            accsre=frappe.db.sql("""select account from `tabExpense Accounts` where parenttype='Broiler Indirect Expenses' and parent='{0}' """.format(re.name),as_dict=1)
            if accsre:
                for ac in accsre:
                    acc.append(ac.account)
                accs='","'.join(acc)
            exppjt=frappe.db.sql(""" select IFNULL(sum(debit_in_account_currency), 0) as debit,IFNULL(sum(credit_in_account_currency), 0) as credit from `tabGL Entry` where is_cancelled!=1 and account in ("{0}") and posting_date between '{1}' and '{2}' and  project='{3}' """.format(accs,start,end,layer.name),as_dict=1,debug=0)
            if exppjt:
                ind_expanse+=float(exppjt[0].debit)-float(exppjt[0].credit)

            expwpjt=frappe.db.sql(""" select IFNULL(sum(debit_in_account_currency), 0) as debit,IFNULL(sum(credit_in_account_currency), 0) as credit from `tabGL Entry` where is_cancelled!=1 and account in ("{0}") and posting_date between '{1}' and '{2}' and project is NULL """.format(accs,start,end),as_dict=1,debug=0)
            if expwpjt:
                exp=float(expwpjt[0].debit)-float(expwpjt[0].credit)
                if exp != 0:
                    if wageper:
                        ind_expanse+=(float(exp)/100)*float(wageper)                
                    else:
                        batchcount=frappe.db.sql(""" select count(name) as cnt from `tabLayer Batch` where  status='Open' """,as_dict=1,debug=0)
                        if batchcount:
                            ind_expanse+=float(exp)/float(batchcount[0].cnt)

            rear_ind_expanse_tot+=ind_expanse
            col_tot+=ind_expanse
            rear_ind_expanse_data.update({'title':re.title,'amt':ind_expanse})
            rear_ind_expanse_item.append(rear_ind_expanse_data)
        rear_ind_expanse.append(rear_ind_expanse_item)
     
    #----------------------------------
    reat_tot.append(col_tot)
    #------------------------------
    rearing_ind_array=[]
    if rear_ind_itm_cnt:
        for num in range(0,rear_ind_itm_cnt):
            i=0
            row_tot=0
            row_ind=[]
            for indx in rear_ind_expanse:                
                if i==0:
                    row_ind.append(indx[num].get('title'))
                    row_ind.append(indx[num].get('amt'))
                    row_tot+=float(indx[num].get('amt'))
                else:
                    row_ind.append(indx[num].get('amt'))
                    row_tot+=float(indx[num].get('amt'))
                i+=1
            #row_ind.append(row_tot)
            rearing_ind_array.append(row_ind)
            
    
    
    #reat_tot_tot=float(rear_doc_tot)+float(rear_feed_tot)+float(rear_vaccine_tot)+float(rear_medicine_tot)+float(rear_bio_tot)+float(rear_miscel_tot)+float(rear_other_tot)+float(rear_wages_tot)+float(rear_ind_expanse_tot)
    #rear_lbl.append('Total')
    #rear_doc.append(rear_doc_tot)
    #rear_feed.append(rear_feed_tot)
    #rear_vaccine.append(rear_vaccine_tot)
    #rear_medicine.append(rear_medicine_tot)
    #rear_bio.append(rear_bio_tot)
    #rear_miscel.append(rear_miscel_tot)
    #rear_other.append(rear_other_tot)
    #rear_wages.append(rear_wages_tot)
    #reat_tot.append(reat_tot_tot)
    rear_graph=[]
    if rear_doc_tot:
        rear_graph.append({"label":"Doc","data":flt(rear_doc_tot,2)})
    if rear_feed_tot:
        rear_graph.append({"label":"Feed","data":flt(rear_feed_tot,2)})
    if rear_vaccine_tot:
        rear_graph.append({"label":"Vaccine","data":flt(rear_vaccine_tot,2)})
    if rear_medicine_tot:
        rear_graph.append({"label":"Medicine","data":flt(rear_medicine_tot,2)})
    if rear_other_tot:
        rear_graph.append({"label":"Other","data":flt(rear_other_tot,2)})
    if rear_wages_tot:
        rear_graph.append({"label":"Wages","data":flt(rear_wages_tot,2)})

#======================================
    date_lbl=rear_start_date.strftime("%d/%m/%y")+'-'+rear_end_date.strftime("%d/%m/%y")
    rear_html+='<tr class="table-secondary">'
    rear_html+='<th scope="col">Expense</th>'
    rear_html+='<th scope="col" colspan="2" class="text-right">'+date_lbl+'</th>'
    rear_html+='</tr>'

    
    #--------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_doc:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+' :(Qty-'+str(layer.doc_placed)+')</th>'
        else:
            rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
    #-----------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_feed:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2" >'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
    #-----------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_vaccine:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_medicine:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    
    #---------------------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_other:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'

     #---------------------------------------------------
    rear_html+='<tr class="table-secondary">'
    i=0
    for doc in rear_direct:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    rear_html+='<tr>'
    i=0
    for doc in rear_wages:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
     #---------------------------------------------------
    rear_html+='<tr class="table-secondary">'
    i=0
    for doc in rear_tot_direct:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
    #-------------------------------------------------
    if len(rearing_ind_array):
        rear_html+='<tr class="table-secondary">'
        i=0
        for lbl in rear_lbl:
            if i==0:
                rear_html+='<th scope="col">Indirect Expense</th>'
            else:
                rear_html+='<td class="text-right" colspan="2"></td>'
            i+=1
        rear_html+='</tr>'
        
        
        for redind in rearing_ind_array:
            rear_html+='<tr >'
            i=0
            for reind in redind:
                if i==0:
                    rear_html+='<th scope="col">'+str(reind)+'</th>'
                else:
                    rear_html+='<td class="text-right" colspan="2">'+str(frappe.utils.fmt_money(reind))+'</td>'
                i+=1
            rear_html+='</tr>'

    #---------------------------------------------------
    rear_html+='<tr class="table-secondary">'
    i=0
    for doc in reat_tot:
        if i==0:
            rear_html+='<th scope="row">'+str(doc)+'</th>'
        else:
            rear_html+='<td class="text-right" colspan="2" >'+str(frappe.utils.fmt_money(doc))+'</td>'
        i+=1

    rear_html+='</tr>'
    #---------------------------------------------------
    #Batch Type
    #break_even_period

    

    #------------------------blank ---------------------------
      
    

#=======================================================================
    

    #Layer Budget
    budget=frappe.db.get_value("Broiler Budget",layer.name,['doc','feed','medicine','vaccine','wages','others','production','sales'],as_dict=1)
    doc=0
    feed=0
    medicine=0
    vaccine=0
    wages=0
    others=0
    production=0
    sales=0
    if budget:
       
        doc=budget.doc
        feed=budget.feed
        medicine=budget.medicine
        vaccine=budget.vaccine
        wages=budget.wages
        others=budget.others
        production=budget.production
        sales=budget.sales
    budget_html='' 
    budget_html+='<tr><td colspan="2"><b>'+str(layer.name)+'<b></td></tr>' 
    budget_html+='<tr><th style="width:30%;">Doc</th><td>'+str(doc)+'</td></tr>'      
    budget_html+='<tr><th>Feed</th><td>'+str(feed)+'</td></tr>'    
    budget_html+='<tr><th>vaccine<td>'+str(vaccine)+'</td></tr>'    
    budget_html+='<tr><th>Medicine</th><td>'+str(medicine)+'</td></tr>'    
    budget_html+='<tr><th>Wages</th><td>'+str(wages)+'</td></tr>'    
    budget_html+='<tr><th>Other</th><td>'+str(others)+'</td></tr>'    
    budget_html+='<tr><th>Production</th><td>'+str(production)+'</td></tr>'    
    budget_html+='<tr><th>Sales</th><td>'+str(sales)+'</td></tr>'
    
    production_cnt=0
    if manuf:
        for man in manuf:
            if man.item_code==finished_product:
                production_cnt=man.qty
                rear_html+='<tr ><td> &nbsp;</td> <td colspan="2">&nbsp;</td> </tr>'
                rear_html+='<tr class="table-secondary"><td> Broiler Produced</td> <td class="text-right"> Qty</td> <td class="text-right"> Amount (With Out Indirect Exp.)</td></tr>'
                rear_html+='<tr><td>'+str(getitem_name(man.item_code))+'</td><td class="text-right">'+str(man.qty)+'</td><td class="text-right">'+str(frappe.utils.fmt_money(man.amount))+'</td></tr>'
                if layer.current_alive_chicks:
                    rear_html+='<tr><td>Live Chicken</td><td class="text-right">'+str(float(layer.current_alive_chicks))+'</td><td class="text-right"></td></tr>'
        
        for man in manuf:
            if man.item_code==cull:
                rear_html+='<tr ><td> &nbsp;</td> <td colspan="2">&nbsp;</td> </tr>'
                rear_html+='<tr class="table-secondary"><td>Mortality / Cull Items</td> <td class="text-right"> Qty</td> <td class="text-right"> Amount</td></tr>'
                if layer.total_mortaliy > man.qty:
                    rte=float(man.amount)/float(man.qty)
                    amt=float(rte)*float(layer.total_mortaliy)
                    rear_html+='<tr><td>'+str(getitem_name(man.item_code))+'</td><td class="text-right">'+str(layer.total_mortaliy)+'</td><td class="text-right">'+str(frappe.utils.fmt_money(amt))+'</td></tr>'
                else:
                    rear_html+='<tr><td>'+str(getitem_name(man.item_code))+'</td><td class="text-right">'+str(man.qty)+'</td><td class="text-right">'+str(frappe.utils.fmt_money(man.amount))+'</td></tr>'
        
    cost=float(col_tot)/(float(production_cnt)+float(layer.current_alive_chicks))
    rear_html+='<tr ><td> &nbsp;</td> <td colspan="2">&nbsp;</td> </tr>'
    rear_html+='<tr class="table-secondary"><td>Cost/Chicken</td> <td class="text-right" colspan="2"><b>'+str(flt(cost,4))+'</b></td></tr>'
    
    curdate=layer.end_date
    if getdate(layer.end_date) > getdate(nowdate()):
        curdate=nowdate()
        
    
    curdate=getdate(curdate).strftime("%d-%m-%Y")
    
    msg='Batch : '+str(layer.name)+'&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Date: '+str(getdate(layer.receiving_date).strftime("%d-%m-%Y"))+' To '+str(curdate)
    return {'rear':rear_html,'budget':budget_html,'rear_graph':rear_graph,'msg':msg}

def getitem_name(item_code):
    return frappe.db.get_value('Item',item_code,'item_name')

#---------------------------------- graphs ------------------------------
@frappe.whitelist()
def get_rear_mor_graph(batch):
    bth=frappe.db.get_value('Broiler Batch',batch,['strain','receiving_date','doc_placed','end_date'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.receiving_date)
    daycnt=date_diff(getdate(bth.end_date),date)+1
   
    #actmort=frappe.db.get_all('Rearing Period Performance',filters={'parent':strain},fields=['age','mortality'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') as dy,IFNULL(total,0) as mort FROM  `tabMortality` 
    WHERE  parent='{0}' and `date` >='{1}' order by date """.format(batch,date),as_dict=1,debug=0)
   
    actmort=[]
    for x in range(daycnt):
        rx={'age':x,'mortality':0,'act_mortality':0}
        for ac in actqry:
            if ac.dy==x:
                mortp=(float(ac.mort)*100)/doc_placed
                rx.update({'act_mortality':mortp})
        actmort.append(rx)

    return {'ideal':actmort}

@frappe.whitelist()
def get_rear_feed_graph(batch):
    bth=frappe.db.get_value('Broiler Batch',batch,['strain','receiving_date','doc_placed','end_date'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.receiving_date)
    daycnt=date_diff(getdate(bth.end_date),date)+1
    stdfeed=frappe.db.get_all('Broiler Period Performance',filters={'parent':strain},fields=['age','daily_feed_intake','cum_feed_intake','cum_feed_conversion'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') as wk,IFNULL(starter_qty,0)+IFNULL(finisher_qty,0) as qty FROM  `tabFeed` 
    WHERE parent='{0}' and `date` >='{1}' order by wk """.format(batch,date),as_dict=1,debug=0)
   
    mortality=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') as wk,IFNULL(total,0) as mort FROM  `tabMortality` 
    WHERE  parent='{0}' and `date` >='{1}' order by wk """.format(batch,date),as_dict=1,debug=0)
    
    actmort=[]
    for x in range(daycnt):
        rx={'age':x,'feed':0,'act_feed':0}

        mort=0
        livechick=0
        for mo in mortality:
            if mo.wk<=x:
                mort+=float(mo.mort)
        
        livechick+=float(doc_placed)-float(mort)

        feed=0
        qty=0
        for ac in actqry:
            if ac.wk<=x:
                qty+=ac.qty
                
        if qty:
            feed=float(qty*50000)/float(livechick)
            rx.update({'act_feed':flt(feed,3)})

        for ac in stdfeed:
            if ac.age==x:               
                rx.update({'feed':ac.cum_feed_intake})

        actmort.append(rx)
    
    return {'ideal':actmort}

@frappe.whitelist()
def get_rear_weight_graph(batch):
    bth=frappe.db.get_value('Broiler Batch',batch,['strain','receiving_date','doc_placed','end_date'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.receiving_date)
    daycnt=date_diff(getdate(bth.end_date),date)+1
    stdweight=frappe.db.get_all('Broiler Period Performance',filters={'parent':strain},fields=['age','weight','daily_gain','avg_daily_gain'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') as wk,IFNULL(weight,0) as weight,IFNULL(lag(weight,1) over(order by `date`),0) as prev_weight FROM  `tabWeight` 
    WHERE parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
   
   
    actmort=[]
    for x in range(daycnt):
        rx={'age':x,'weight':0,'act_weight':0}

        for ac in stdweight:
            if ac.age==x:               
                rx.update({'weight':ac.weight})
        wgain=0
        for ac in actqry:
            if ac.wk==x:
                #wgain=ac.weight-ac.prev_weight
                rx.update({'act_weight':ac.weight})
    
        actmort.append(rx)
        
    return {'ideal':actmort}

    
@frappe.whitelist()
def get_frc_graph(batch):
    bth=frappe.db.get_value('Broiler Batch',batch,['strain','receiving_date','doc_placed','end_date'],as_dict=1)
    strain=bth.strain
    doc_placed=bth.doc_placed
    date=getdate(bth.receiving_date)
    daycnt=date_diff(getdate(bth.end_date),date)+1
    stdfcr=frappe.db.get_all('Broiler Period Performance',filters={'parent':strain},fields=['age','fcr'],order_by='age')
    
    actqry=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') as wk,IFNULL(starter_qty,0)+IFNULL(finisher_qty,0) as qty FROM  `tabFeed` 
    WHERE parent='{0}' and `date` >='{1}' order by wk """.format(batch,date),as_dict=1,debug=0)
   
    mortality=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') as wk,IFNULL(total,0) as mort FROM  `tabMortality` 
    WHERE  parent='{0}' and `date` >='{1}' order by wk """.format(batch,date),as_dict=1,debug=0)

    weight=frappe.db.sql(""" SELECT DATEDIFF(`date`,'{1}') as wk,IFNULL(weight,0) as weight,IFNULL(lag(weight,1) over(order by `date`),0) as prev_weight FROM  `tabWeight` 
    WHERE parent='{0}' and `date` >='{1}' GROUP BY wk order by wk """.format(batch,date),as_dict=1,debug=0)
    
    actmort=[]
    for x in range(daycnt):
        rx={'age':x,'fcr':0,'act_fcr':0}

        for ac in stdfcr:
            if ac.age==x:               
                rx.update({'fcr':ac.fcr})

        mort=0
        livechick=0
        for mo in mortality:
            if mo.wk<=x:
                mort+=float(mo.mort)
        
        livechick+=float(doc_placed)-float(mort)

        feed=0
        qty=0
        for ac in actqry:
            if ac.wk<=x:
                qty+=ac.qty
                
        if qty:
            feed=float(qty*50000)/float(livechick)

        for we in weight:
            fcr=0
            if we.wk==x:
                if feed and we.weight:
                    fcr=feed/we.weight
                rx.update({'act_fcr':flt(fcr,3)})

        
    
        actmort.append(rx)
    return {'ideal':actmort}
#-----------------------------------------------------------------------------------
import os
#import tempfile


#from frappe.utils.response import download_file
@frappe.whitelist()
def down_report(company,batch,rearing=None,budget=None,rearing_gp=None,rearing_mor_gp=None,rearing_feed_gp=None,rearing_weight_gp=None,frc_gp=None):
    import json
    import pandas as pd
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook    
    from openpyxl.chart import (
        PieChart, LineChart,
        ProjectedPieChart,
        Reference
    )
    
    from openpyxl.styles import Font
    from openpyxl.chart.axis import DateAxis
    #from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Border, Side

    ft = Font(bold=True)

    rearingary=json.loads(rearing)
    budgetary=json.loads(budget)
    gp_rear=json.loads(rearing_gp)
    gp_rear_mor=json.loads(rearing_mor_gp)   
    gp_rearing_feed=json.loads(rearing_feed_gp)   
    gp_rearing_weight=json.loads(rearing_weight_gp)
    gp_laying_performance=json.loads(frc_gp)

    rrowlen=len(rearingary)
    rcollen=len(rearingary[0])
    browlen=len(budgetary)
    bcollen=len(budgetary[0])
    report=[]
    
    rear_mor=[]
    lbl=['Age','Std. Mortality %','Actual Mortality %']
    #report.append(['Age','Std. Mortality %','Actual Mortality %','Age','Std. Min Feed','Std. Max Feed','Actual Feed','Age','Std. Min Weigt','Std. Max Weight','Actual Weight','Age','Std. Min Eggs','Std. Max Eggs','Actual Eggs'])

    rear_mor.append(lbl)
    j=0
    if len(gp_rear_mor):        
        for rm in gp_rear_mor:
            rear_mor.append([rm.get('age'),rm.get('mortality'),rm.get('act_mortality')])            
            #report.append([rm.get('age'),rm.get('mortality'),rm.get('act_mortality'),rm.get('age'),gp_rearing_feed[0][j].get('v1'),gp_rearing_feed[0][j].get('v2'),gp_rearing_feed[0][j].get('act_feed'),rm.get('age'),gp_rearing_weight[0][j].get('v1'),gp_rearing_weight[0][j].get('v2'),gp_rearing_weight[0][j].get('act_weight'),rm.get('age'),0,0,0])
            
            j+=1
            
    rear_exp=[]
    lbl=['','Expenses','Totals']
    rear_exp.append(lbl)
    if len(gp_rear):        
        for rm in gp_rear:
            rear_exp.append(['',rm.get('label'),rm.get('data')])            
          
    gprearing_feed=[]
    gprearing_feed.append(['Age','Std Feed','Actual Feed'])
    #frappe.msgprint(str(gp_rearing_feed))
    if len(gp_rearing_feed):        
        for lm in gp_rearing_feed:
            gprearing_feed.append([lm.get('age'),float(lm.get('feed')),float(lm.get('act_feed'))])

    
    gprearing_weight=[]
    gprearing_weight.append(['Age','Std Weigt','Actual Weight'])
    if len(gp_rearing_weight):        
        for lm in gp_rearing_weight:
            gprearing_weight.append([lm.get('age'),float(lm.get('weight')),float(lm.get('act_weight'))])

    
    gplaying_performance=[]
    gplaying_performance.append(['Age','Std. FCR','Actual FCR'])
    if len(gp_laying_performance):        
        for lm in gp_laying_performance:
            gplaying_performance.append([lm.get('age'),float(lm.get('fcr')),float(lm.get('act_fcr'))])

    wb = Workbook()
    ws = wb.active
    ws.title = "Rearing"    
    ws1 = wb.create_sheet("Budget")   
    ws5 = wb.create_sheet("Mor. GPH")    
    ws7 = wb.create_sheet("Feed. GPH")    
    ws9 = wb.create_sheet("Weight. GPH")
    ws11 = wb.create_sheet("FCR")
    

    for row in rearingary:
       ws.append(row)

    for row in budgetary:
       ws1.append(row)

    ws.append([''])
    for row in rear_exp:
        ws.append(row)
               
    for row in rear_mor:
        ws5.append(row)

    for row in gprearing_feed:
        ws7.append(row)
    
    for row in gprearing_weight:
        ws9.append(row)

    for row in gplaying_performance:
        ws11.append(row)


    #for row in report:
     #   rep.append(row)
    
    yellow = "00D5D7D9"
    black="00000000"
    thin = Side(border_style="thin", color=black)
    double = Side(border_style="double", color=black)
    thick = Side(border_style="thick", color=black)

    #===========================================================
    rlbl=getColumnName(rcollen)
    rhd="A1:"+str(rlbl)+str(1)
    
    for row in ws[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:"+str(rlbl)+str(rrowlen)
    for row in ws[rhd]:
        rowhed=0
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if "Total" in str(cell.value):
                rowhed=1
            if rowhed==1:
                cell.font = ft
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:A"+str(rrowlen)
    for row in ws[rhd]:
        for cell in row:
            #cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    #---------------------------------------------------------------------------------
    rlbl=getColumnName(bcollen)
    rhd="A1:"+str(rlbl)+str(1)
    
    for row in ws1[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:"+str(rlbl)+str(browlen)
    for row in ws1[rhd]:
        rowhed=0
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if "Total" in str(cell.value):
                rowhed=1
            if rowhed==1:
                cell.font = ft
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")

    rhd="A1:A"+str(browlen)
    for row in ws1[rhd]:
        for cell in row:
            cell.font = ft
            cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
    #----------------------------------------------------------------------------------------------
    

    if len(gp_rear):
        maxr=len(gp_rear)+1
        rhd="B"+str(rrowlen+2)
        pie = PieChart()
        #labels = Reference(ws3, min_col=2, min_row=2, max_row=maxr)
        labels = "'Rearing'!$B$"+str(rrowlen+3)+":$B$"+str(rrowlen+1+maxr)
        #frappe.msgprint(str(labels))
        #data = Reference(ws3, min_col=3, min_row=1, max_row=maxr)
        data = "'Rearing'!$C$"+str(rrowlen+3)+":$C$"+str(rrowlen+1+maxr)
        #frappe.msgprint(str(data))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Rearing"
        ws.add_chart(pie, rhd)
        
    
    from copy import deepcopy
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import RichTextProperties

    if len(gp_rear_mor):
        
        maxr=len(gp_rear_mor)+1
        c2 = LineChart()
        c2.title = "Mortality"
        c2.style = 13
        c2.x_axis.title = 'Days'
        c2.y_axis.title = 'Mortality %'
        c2.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c2.x_axis.delete = False
        c2.y_axis.delete = False
        c2.height = 10 # default is 7.5
        c2.width = 25 # default is 15
        data = Reference(ws5, min_col=2, min_row=1, max_col=3, max_row=maxr)
        #frappe.msgprint(str(data)) #= 'Rear. Mor. GPH'!$B$1:$C$18
        c2.add_data(data, titles_from_data=True)
        dates = Reference(ws5, min_col=1, min_row=2, max_row=maxr)
        #frappe.msgprint(str(dates)) #='Rear. Mor. GPH'!$A$2:$A$18
        c2.set_categories(dates)
        s2 = c2.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c2.series[1]
        s2.graphicalProperties.line.solidFill = "2490ef"
        s2.graphicalProperties.line.width = 30000 # width in EMUs
        ws5.add_chart(c2, "A1")

        #=========================

        
    #----------------------------------------------------

    if len(gprearing_feed):
        
        maxr=len(gprearing_feed)
        c3 = LineChart()
        c3.title = "Feed"
        c3.style = 13
        c3.x_axis.title = 'Days'
        c3.y_axis.title = 'Cum. Feed Intake'
        c3.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c3.x_axis.delete = False
        c3.y_axis.delete = False
        c3.height = 10 # default is 7.5
        c3.width = 45 # default is 15
        data = Reference(ws7, min_col=2, min_row=1, max_col=3, max_row=maxr)            
        c3.add_data(data, titles_from_data=True)
        dates = Reference(ws7, min_col=1, min_row=2, max_row=maxr)        
        c3.set_categories(dates)
        s2 = c3.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c3.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        #s2 = c3.series[2]
        #s2.graphicalProperties.line.solidFill = "2490ef"
        #s2.graphicalProperties.line.width = 30000 # width in EMUs        
       

        ws7.add_chart(c3, "A1")

        #=========================
    

    #----------------------------------------------------

    if len(gprearing_weight):
        
        maxr=len(gprearing_weight)
        c3 = LineChart()
        c3.title = "Weight"
        c3.style = 13
        c3.x_axis.title = 'Days'
        c3.y_axis.title = 'Weight Gain'
        c3.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c3.x_axis.delete = False
        c3.y_axis.delete = False
        c3.height = 10 # default is 7.5
        c3.width = 40 # default is 15
        data = Reference(ws9, min_col=2, min_row=1, max_col=3, max_row=maxr)            
        c3.add_data(data, titles_from_data=True)
        dates = Reference(ws9, min_col=1, min_row=2, max_row=maxr)        
        c3.set_categories(dates)
        s2 = c3.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c3.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        #s2 = c3.series[2]
        #s2.graphicalProperties.line.solidFill = "2490ef"
        #s2.graphicalProperties.line.width = 30000 # width in EMUs        
       

        ws9.add_chart(c3, "A1")

    

    if len(gplaying_performance):
        maxr=len(gplaying_performance)
        rhd="A1"
        
        c4 = LineChart()
        c4.title = "Performance"
        c4.style = 13
        c4.x_axis.title = 'Days'
        c4.y_axis.title = 'FCR'
        c4.x_axis.title.txPr = RichText(bodyPr=RichTextProperties(rot="-180"))
        c4.x_axis.delete = False
        c4.y_axis.delete = False
        c4.height = 10 # default is 7.5
        c4.width = 40 # default is 15
        data = Reference(ws11, min_col=2, min_row=1, max_col=3, max_row=maxr)
        c4.add_data(data, titles_from_data=True)
        dates = Reference(ws11, min_col=1, min_row=2, max_row=maxr)
        c4.set_categories(dates)
        s2 = c4.series[0]
        s2.graphicalProperties.line.solidFill = "ff2e2e"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        s2 = c4.series[1]
        s2.graphicalProperties.line.solidFill = "55af46"
        s2.graphicalProperties.line.width = 30000 # width in EMUs

        #s2 = c4.series[2]
        #s2.graphicalProperties.line.solidFill = "2490ef"
        #s2.graphicalProperties.line.width = 30000 # width in EMUs 2490ef

        ws11.add_chart(c4, rhd)

    #file_name = 'broiler_dash.xlsx'
    file_name=batch+'.xlsx'    
    temp_file=os.path.join(frappe.utils.get_bench_path(), "logs", file_name)
    wb.save(temp_file)
    return temp_file

def getColumnName(n):
 
    # initialize output string as empty
    result = ''
 
    while n > 0:
 
        # find the index of the next letter and concatenate the letter
        # to the solution
 
        # here index 0 corresponds to 'A', and 25 corresponds to 'Z'
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26
 
    return result[::-1]


@frappe.whitelist()
def down_file(file=None):
    from frappe.utils.file_manager import download_file
    file_name = os.path.basename(file)
    with open(file, "rb") as fileobj:
        filedata = fileobj.read()
    frappe.response['content_type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    frappe.response['content_disposition'] = 'attachment; filename="{0}"'.format(file_name)
    frappe.local.response.filename = file_name
    frappe.local.response.filecontent = filedata
    frappe.local.response.type = "download"

